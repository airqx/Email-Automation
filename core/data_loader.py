"""
Excel and CSV file reading and column auto-detection.
"""

import csv
import os
import sys
from typing import List, Tuple, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Please run: pip3 install openpyxl")
    sys.exit(1)

# Keywords for auto-detecting the email column (English + Arabic).
EMAIL_KEYWORDS = [
    "email", "e-mail", "mail",
    "البريد", "البريد الالكتروني", "البريد الإلكتروني",
]

# Keywords for auto-detecting the name column (English + Arabic).
NAME_KEYWORDS = [
    "name", "full name", "first name", "firstname",
    "الاسم", "اسم", "الاسم الكامل",
]

# Keywords for auto-detecting the attachment path column (English + Arabic).
ATTACHMENT_PATH_KEYWORDS = [
    "attachment", "attachment path", "file", "file path",
    "pdf", "image", "document", "doc", "doc path",
    "path", "attachment_file", "attachment_path",
    "مرفق", "المرفق", "مسار", "مسار المرفق", "ملف", "مسار الملف",
]


def read_csv_file(file_path: str) -> Tuple[List[str], List[dict]]:
    """
    Read a CSV file and return its headers and data rows.

    The first row is treated as headers.
    All subsequent non-empty rows are returned as a list of dicts keyed by header.

    Parameters:
        file_path (str): Path to the .csv file.

    Returns:
        Tuple[List[str], List[dict]]: (headers, rows).

    Side effects:
        Exits the program if the file cannot be read.
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            headers = reader.fieldnames

            if not headers:
                print("ERROR: CSV file is empty or has no headers.")
                sys.exit(1)

            headers = [str(h).strip() for h in headers]

            data = []
            for row in reader:
                # Skip empty rows
                if not any(row.values()):
                    continue

                # Clean up values
                row_dict = {
                    k: str(v).strip() if v is not None else ""
                    for k, v in row.items()
                }
                data.append(row_dict)

            return headers, data

    except Exception as e:
        print(f"ERROR: Could not read CSV file '{file_path}': {e}")
        sys.exit(1)


def read_excel_file(file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[dict]]:
    """
    Read an Excel (.xlsx) file and return its headers and data rows.

    Opens the workbook in read-only mode. The first row is treated as headers.
    All subsequent non-empty rows are returned as a list of dicts keyed by header.

    Parameters:
        file_path (str): Path to the .xlsx file.
        sheet_name (Optional[str]): Specific sheet to read. None = active sheet.

    Returns:
        Tuple[List[str], List[dict]]: (headers, rows).

    Side effects:
        Exits the program if the file cannot be read or the sheet is not found.
    """
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"ERROR: Could not read Excel file '{file_path}': {e}")
        sys.exit(1)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found in workbook.")
            print(f"Available sheets: {', '.join(workbook.sheetnames)}")
            sys.exit(1)
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)

    headers = next(rows_iter, None)
    if not headers:
        print("ERROR: Excel file is empty or has no headers.")
        sys.exit(1)

    headers = [
        str(h).strip() if h is not None else f"Column_{i}"
        for i, h in enumerate(headers)
    ]

    data = []
    for row in rows_iter:
        if not any(row):
            continue

        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = str(value).strip() if value is not None else ""
        data.append(row_dict)

    return headers, data


def read_data_file(file_path: str, sheet_name: Optional[str] = None) -> Tuple[List[str], List[dict]]:
    """
    Read a data file (CSV or Excel) and return its headers and data rows.

    Automatically detects the file type based on the file extension.
    Supports: .csv, .xlsx, .xls

    Parameters:
        file_path (str): Path to the data file.
        sheet_name (Optional[str]): Specific sheet to read (Excel only). None = active sheet.

    Returns:
        Tuple[List[str], List[dict]]: (headers, rows).

    Side effects:
        Exits the program if the file format is unsupported or cannot be read.
    """
    _, ext = os.path.splitext(file_path.lower())

    if ext == '.csv':
        if sheet_name:
            print("WARNING: sheet_name parameter is ignored for CSV files.")
        return read_csv_file(file_path)
    elif ext in ['.xlsx', '.xls']:
        return read_excel_file(file_path, sheet_name)
    else:
        print(f"ERROR: Unsupported file format '{ext}'.")
        print("Supported formats: .csv, .xlsx, .xls")
        sys.exit(1)


def detect_column(headers: List[str], keywords: List[str]) -> Optional[str]:
    """
    Auto-detect a column name by matching headers against known keywords.

    Performs a case-insensitive substring search. Returns the first match.

    Parameters:
        headers (List[str]): Column header names from the Excel file.
        keywords (List[str]): Keywords to search for.

    Returns:
        Optional[str]: The original header name if matched, None otherwise.
    """
    headers_lower = {h.lower().strip(): h for h in headers}

    for keyword in keywords:
        keyword_lower = keyword.lower()
        for header_lower, original_header in headers_lower.items():
            if keyword_lower in header_lower:
                return original_header
    return None


def detect_email_column(headers: List[str]) -> Optional[str]:
    """
    Auto-detect the email column from the available headers.

    Parameters:
        headers (List[str]): Column header names.

    Returns:
        Optional[str]: The matched email column name, or None if not found.
    """
    return detect_column(headers, EMAIL_KEYWORDS)


def detect_name_column(headers: List[str]) -> Optional[str]:
    """
    Auto-detect the name column from the available headers.

    Parameters:
        headers (List[str]): Column header names.

    Returns:
        Optional[str]: The matched name column name, or None if not found.
    """
    return detect_column(headers, NAME_KEYWORDS)


def detect_attachment_path_column(headers: List[str]) -> Optional[str]:
    """
    Auto-detect the attachment path column from the available headers.

    Parameters:
        headers (List[str]): Column header names.

    Returns:
        Optional[str]: The matched attachment path column name, or None if not found.
    """
    return detect_column(headers, ATTACHMENT_PATH_KEYWORDS)


def get_column_interactive(headers: List[str], column_type: str) -> str:
    """
    Prompt the user to manually select a column name from the available headers.

    Parameters:
        headers (List[str]): Available column header names.
        column_type (str): Label for the column (e.g., "EMAIL", "NAME").

    Returns:
        str: The exact column name selected by the user.
    """
    print(f"\nAvailable columns in the data file:")
    for i, header in enumerate(headers, 1):
        print(f"  {i}. {header}")

    while True:
        user_input = input(f"\nEnter the exact column name for {column_type}: ").strip()
        if user_input in headers:
            return user_input
        print(f"ERROR: '{user_input}' is not a valid column name. Please try again.")